import sqlite3
import pandas as pd
import streamlit as st
import io
import os
from datetime import datetime
import pytz
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

# --- 1. 데이터베이스 초기화 함수 ---
def init_db():
    conn = sqlite3.connect("kostal_rework_v2.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            author TEXT,
            item_name TEXT,
            is_update TEXT,  
            is_dtc TEXT       
        )
    """
    )
    conn.commit()
    return conn

conn = init_db()

# --- 2. 현재 한국 시간(KST)을 구하는 함수 ---
def get_current_kst_time():
    kst = pytz.timezone('Asia/Seoul')
    now = datetime.now(kst)
    return now.strftime('%Y-%m-%d %H:%M')

# --- 3. 데이터베이스 조작 함수들 ---
def insert_data(author, item_name, is_update, is_dtc):
    current_time = get_current_kst_time()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO items (timestamp, author, item_name, is_update, is_dtc) VALUES (?, ?, ?, ?, ?)",
        (current_time, author, item_name, is_update, is_dtc),
    )
    conn.commit()

def update_data(item_id, author, item_name, is_update, is_dtc):
    current_time = get_current_kst_time()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE items 
        SET timestamp = ?, author = ?, item_name = ?, is_update = ?, is_dtc = ? 
        WHERE id = ?
        """,
        (current_time, author, item_name, is_update, is_dtc, item_id),
    )
    conn.commit()

def delete_data(item_id):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM items WHERE id = ?", (item_id,))
    conn.commit()

def load_data():
    return pd.read_sql_query("SELECT * FROM items ORDER BY id DESC", conn)

def clear_data():
    cursor = conn.cursor()
    cursor.execute("DELETE FROM items")
    conn.commit()

# --- 4. 엑셀 파일 내에 사진을 품목명 시트로 저장하는 함수 ---
def save_image_to_excel(item_name, is_update, is_dtc, uploaded_file):
    excel_filename = "KOSTAL_photo_registry.xlsx"
    
    if os.path.exists(excel_filename):
        wb = openpyxl.load_workbook(excel_filename)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "임시_기본시트"
        
    clean_sheet_name = "".join(c for c in item_name if c not in r'\/??*[]:').strip()[:30]
    if not clean_sheet_name:
        clean_sheet_name = "정제된_품목명"
        
    if clean_sheet_name in wb.sheetnames:
        ws = wb[clean_sheet_name]
    else:
        ws = wb.create_sheet(title=clean_sheet_name)
        ws['A1'] = "등록 시간"
        ws['B1'] = "품목명"
        ws['C1'] = "업데이트 여부"
        ws['D1'] = "DTC 여부"
        ws['E1'] = "첨부 사진"
        
    image = Image.open(uploaded_file)
    
    # 📱 모달/모바일 폰카 대용량 사진 대응 리사이징 (가로 400px 제한 및 해상도 최적화)
    image.thumbnail((400, 400))
    
    img_byte_arr = io.BytesIO()
    # JPEG 포맷 + 용량 최적화용 quality 스케일링 적용 (용량 최소화)
    if image.mode in ("RGBA", "P"):
        image = image.convert("RGB")
    image.save(img_byte_arr, format='JPEG', quality=75, optimize=True)
    img_byte_arr.seek(0)
    
    xl_img = OpenpyxlImage(img_byte_arr)
    
    next_row = ws.max_row + 1 if ws.max_row > 1 or ws['A1'].value != "등록 시간" else 2
    
    ws[f'A{next_row}'] = get_current_kst_time()
    ws[f'B{next_row}'] = item_name
    ws[f'C{next_row}'] = is_update
    ws[f'D{next_row}'] = is_dtc
    
    ws.add_image(xl_img, f'E{next_row}')
    ws.row_dimensions[next_row].height = 160  # 행 높이를 사진 비율에 맞춰 소폭 상향
    ws.column_dimensions['E'].width = 45
    
    if "임시_기본시트" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["임시_기본시text"]
        
    wb.save(excel_filename)
    return excel_filename


# --- 5. 모달 팝업(Dialog) 정의 창 ---
@st.dialog("⚠️ 데이터 삭제 확인")
def confirm_delete_dialog(item_id, item_name):
    st.write(f"정말로 품목명 **[{item_name}]** 항목을 삭제하시겠습니까?")
    st.write("삭제된 데이터는 복구할 수 없습니다.")
    st.write("")
    col_pop1, col_pop2 = st.columns(2)
    with col_pop1:
        if st.button("🔴 예, 삭제합니다", use_container_width=True):
            delete_data(item_id)
            st.toast("성공적으로 삭제되었습니다.")
            st.rerun()
    with col_pop2:
        if st.button("취소", use_container_width=True):
            st.rerun()

@st.dialog("🚨 입력 오류 알림")
def validation_error_dialog(message):
    st.error(message)
    st.write("마감 및 식별을 위해 품목명 칸에는 **반드시 숫자 6자리**를 정확히 입력해 주세요.")
    if st.button("확인", use_container_width=True):
        st.rerun()


# --- 6. 수정/삭제 모드 관리를 위한 세션 상태 설정 ---
if "edit_mode" not in st.session_state:
    st.session_state.edit_mode = False
if "edit_id" not in st.session_state:
    st.session_state.edit_id = None
if "form_author" not in st.session_state:
    st.session_state.form_author = ""
if "form_item_name" not in st.session_state:
    st.session_state.form_item_name = ""
if "form_update" not in st.session_state:
    st.session_state.form_update = False
if "form_dtc" not in st.session_state:
    st.session_state.form_dtc = False

# --- 7. Streamlit 웹 UI 구성 ---
st.set_page_config(page_title="KOSTAL Rework System", layout="wide")

st.title("📊 KOSTAL 통합 관리 시스템")
st.write("실시간 리스트 취합 및 품목별 사진 대장 관리 기능이 통합되었습니다.")

# --- 사이드바: 데이터 입력 및 수정 폼 ---
if st.session_state.edit_mode:
    st.sidebar.header("✏️ 데이터 수정 중")
    st.sidebar.info(f"ID 번호 [{st.session_state.edit_id}]번 항목을 수정하고 있습니다.")
else:
    st.sidebar.header("✍️ 새 데이터 입력")

author = st.sidebar.text_input("작성자(이름/부서)", value=st.session_state.form_author)
item_name = st.sidebar.text_input("품목명 (숫자 6자리 입력)", value=st.session_state.form_item_name)

st.sidebar.markdown("**🛠️ 체크 사항 선택**")
chk_update = st.sidebar.checkbox("🔄 업데이트 포함", value=st.session_state.form_update)
chk_dtc = st.sidebar.checkbox("⚠️ DTC 확인 완료", value=st.session_state.form_dtc)

val_update = "Y" if chk_update else "N"
val_dtc = "Y" if chk_dtc else "N"

# 📸 모바일 카메라 실시간 연동 팁 안내 문구 추가
st.sidebar.markdown("---")
st.sidebar.caption("💡 스마트폰 접속 시 아래 버튼을 누르면 즉시 카메라로 실시간 촬영 및 첨부가 가능합니다.")
uploaded_file = st.sidebar.file_uploader("📸 증빙 사진 첨부 (선택 사항)", type=["png", "jpg", "jpeg"])

if st.session_state.edit_mode:
    col_btn1, col_btn2 = st.sidebar.columns(2)
    with col_btn1:
        if st.button("✅ 수정 완료", use_container_width=True):
            if author and item_name:
                if not item_name.isdigit() or len(item_name) != 6:
                    validation_error_dialog(f"입력된 품목명 [{item_name}]은 올바른 형식이 아닙니다.")
                else:
                    update_data(st.session_state.edit_id, author, item_name, val_update, val_dtc)
                    if uploaded_file is not None:
                        save_image_to_excel(item_name, val_update, val_dtc, uploaded_file)
                    st.toast("성공적으로 수정되었습니다!")
                    st.session_state.edit_mode = False
                    st.session_state.edit_id = None
                    st.session_state.form_author = ""
                    st.session_state.form_item_name = ""
                    st.session_state.form_update = False
                    st.session_state.form_dtc = False
                    st.rerun()
            else:
                st.sidebar.error("작성자와 품목명은 필수입니다.")
    with col_btn2:
        if st.button("❌ 취소", use_container_width=True):
            st.session_state.edit_mode = False
            st.session_state.edit_id = None
            st.session_state.form_author = ""
            st.session_state.form_item_name = ""
            st.session_state.form_update = False
            st.session_state.form_dtc = False
            st.rerun()
else:
    if st.sidebar.button("➕ 리스트에 추가", use_container_width=True):
        if author and item_name:
            if not item_name.isdigit() or len(item_name) != 6:
                validation_error_dialog(f"입력된 품목명 [{item_name}]은 올바른 형식이 아닙니다.")
            else:
                insert_data(author, item_name, val_update, val_dtc)
                if uploaded_file is not None:
                    save_image_to_excel(item_name, val_update, val_dtc, uploaded_file)
                st.toast(f"'{item_name}' 항목이 추가되었습니다!")
                st.rerun()
        else:
            st.sidebar.error("작성자와 품목명은 필수 입력 항목입니다.")

# --- 8. 데이터 표시 및 제어 기능 구현 ---
df = load_data()

col1, col2 = st.columns([4, 1])

with col1:
    st.subheader("📋 현재 병합된 리스트 (실시간 업데이트)")
    search_query = st.text_input("🔍 수정/삭제할 품목명 또는 작성자를 입력하세요 (실시간 필터링)", "")
    
    if not df.empty:
        if search_query:
            filtered_df = df[
                df['item_name'].str.contains(search_query, case=False, na=False) | 
                df['author'].str.contains(search_query, case=False, na=False)
            ]
        else:
            filtered_df = df

        if filtered_df.empty:
            st.warning(f"'{search_query}'에 매칭되는 검색 결과가 없습니다.")
        else:
            cols = st.columns([0.6, 0.6, 2.0, 1.5, 2.5, 1.0, 1.0])
            cols[0].write("**수정**")
            cols[1].write("**삭제**")
            cols[2].write("**기록 시간**")
            cols[3].write("**작성자**")
            cols[4].write("**품목명**")
            cols[5].write("**업데이트**")
            cols[6].write("**DTC**")
            st.write("---")
            
            for index, row in filtered_df.iterrows():
                row_cols = st.columns([0.6, 0.6, 2.0, 1.5, 2.5, 1.0, 1.0])
                
                if row_cols[0].button("📝", key=f"edit_{row['id']}", help="수정"):
                    st.session_state.edit_mode = True
                    st.session_state.edit_id = row['id']
                    st.session_state.form_author = row['author']
                    st.session_state.form_item_name = row['item_name']
                    st.session_state.form_update = True if row['is_update'] == "Y" else False
                    st.session_state.form_dtc = True if row['is_dtc'] == "Y" else False
                    st.rerun()
                    
                if row_cols[1].button("🗑️", key=f"del_{row['id']}", help="삭제"):
                    confirm_delete_dialog(row['id'], row['item_name'])
                    
                row_cols[2].write(row['timestamp'])
                row_cols[3].write(row['author'])
                row_cols[4].write(row['item_name'])
                row_cols[5].write(row['is_update'])
                row_cols[6].write(row['is_dtc'])
                st.write(" ") 
    else:
        st.info("현재 입력된 데이터가 없습니다. 왼쪽에서 데이터를 입력해 주세요.")

with col2:
    st.subheader("💾 데이터 다운로드")
    
    if not df.empty:
        towrite = io.BytesIO()
        df.to_excel(towrite, index=False, engine="openpyxl")
        towrite.seek(0)
        st.download_button(
            label="📈 일반 리스트 내보내기",
            data=towrite,
            file_name="KOSTAL_rework_list.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    if os.path.exists("KOSTAL_photo_registry.xlsx"):
        with open("KOSTAL_photo_registry.xlsx", "rb") as f:
            st.download_button(
                label="📸 사진대장 엑셀 내보내기",
                data=f,
                file_name="KOSTAL_photo_registry.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.write("---")
    if st.button("🔴 전체 데이터 초기화", use_container_width=True):
        clear_data()
        if os.path.exists("KOSTAL_photo_registry.xlsx"):
            os.remove("KOSTAL_photo_registry.xlsx")
        st.session_state.edit_mode = False
        st.rerun()